"""
Export Service Module
Generates Excel and PDF reports for developer performance data.
"""

import io
import json
from datetime import datetime, date
from typing import Optional, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.graphics.shapes import Drawing, Line
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt


class ExportService:
    """Service for generating Excel and PDF export reports."""

    # Color definitions
    GREEN_FILL = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    RED_FILL = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    GRAY_FILL = PatternFill(start_color="e9ecef", end_color="e9ecef", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    @staticmethod
    def get_quality_grade(churn_score: float) -> str:
        """Convert churn_score to letter grade. Lower churn = better grade."""
        if churn_score is None:
            return "N/A"
        if churn_score <= 0.2:
            return "A"
        elif churn_score <= 0.4:
            return "B"
        elif churn_score <= 0.6:
            return "C"
        elif churn_score <= 0.8:
            return "D"
        else:
            return "F"

    @staticmethod
    def format_time_hhmm(seconds: int) -> str:
        """Format seconds to HH:MM string."""
        if not seconds:
            return "00:00"
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        return f"{hours:02d}:{minutes:02d}"

    @staticmethod
    def safe_float(value, default=0.0) -> float:
        """Safely convert value to float."""
        try:
            return float(value) if value is not None else default
        except (TypeError, ValueError):
            return default

    @staticmethod
    def safe_int(value, default=0) -> int:
        """Safely convert value to int."""
        try:
            return int(value) if value is not None else default
        except (TypeError, ValueError):
            return default

    def generate_excel(
        self,
        metrics_data: List[dict],
        developers_map: dict,
        from_date: str,
        to_date: str,
        include_charts: bool = True,
        include_summary: bool = True,
        group_by_developer: bool = False,
        include_raw_wakatime: bool = False
    ) -> io.BytesIO:
        """Generate Excel report with multiple sheets."""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Sheet 1: Summary
        if include_summary:
            self._create_summary_sheet(wb, metrics_data, developers_map, from_date, to_date, include_charts)
        
        # Sheet 2: Detailed Data
        self._create_detailed_sheet(wb, metrics_data, developers_map, group_by_developer)
        
        # Sheet 3: Charts (if enabled and has data)
        if include_charts and metrics_data:
            self._create_charts_sheet(wb, metrics_data, developers_map)
        
        # Sheet 4: WakaTime Details (if enabled)
        if include_raw_wakatime:
            self._create_wakatime_sheet(wb, metrics_data, developers_map)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _create_summary_sheet(self, wb: Workbook, data: List[dict], devs: dict, from_date: str, to_date: str, include_charts: bool):
        """Create the summary sheet with statistics."""
        ws = wb.create_sheet("Report Summary")
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "Performance Report"
        ws['A1'].font = Font(size=18, bold=True, color="3b82f6")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Report info
        ws['A3'] = "Date Range:"
        ws['B3'] = f"{from_date} to {to_date}"
        ws['A4'] = "Developers:"
        ws['B4'] = ", ".join([devs.get(d.get('developer_id'), 'Unknown') for d in data[:5]]) + ("..." if len(data) > 5 else "")
        ws['A5'] = "Generated:"
        ws['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Summary Statistics
        ws['A7'] = "SUMMARY STATISTICS"
        ws['A7'].font = Font(size=14, bold=True)
        
        if data:
            total_days = len(data)
            active_days = len([d for d in data if d.get('commits_count', 0) > 0 or d.get('coding_time_seconds', 0) > 0])
            total_commits = sum(d.get('commits_count', 0) for d in data)
            total_lines = sum((d.get('lines_added', 0) or 0) + (d.get('lines_deleted', 0) or 0) for d in data)
            avg_score = sum(d.get('score', 0) or 0 for d in data) / max(total_days, 1)
            avg_coding_hours = sum(d.get('coding_time_seconds', 0) or 0 for d in data) / max(total_days, 1) / 3600
            
            stats = [
                ("Total Records", total_days),
                ("Active Days", f"{active_days} ({active_days/max(total_days,1)*100:.1f}%)"),
                ("Total Commits", total_commits),
                ("Total Lines Changed", f"{total_lines:,}"),
                ("Average Score", f"{avg_score:.1f}"),
                ("Average Coding Hours", f"{avg_coding_hours:.1f}h"),
            ]
            
            for i, (label, value) in enumerate(stats, start=9):
                ws[f'A{i}'] = label
                ws[f'A{i}'].font = Font(bold=True)
                ws[f'B{i}'] = str(value)
        
        # Auto-size columns
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20

    def _create_detailed_sheet(self, wb: Workbook, data: List[dict], devs: dict, group_by_developer: bool):
        """Create the detailed data sheet."""
        ws = wb.create_sheet("Detailed Data")
        
        # Headers
        headers = [
            "Developer Name", "Date", "Commits", "Lines Added", "Lines Deleted",
            "Lines Changed", "Files Modified", "Churn Score", "Coding Time (HH:MM)",
            "Coding Hours", "Active Coding (min)", "Deep Work (min)", "Project Focus %",
            "Context Switches", "Code Reviews", "Start Time", "End Time",
            "Performance Score", "Quality Grade", "Updated At"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = self.THIN_BORDER
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Sort data if grouping
        sorted_data = data
        if group_by_developer:
            sorted_data = sorted(data, key=lambda x: (devs.get(x.get('developer_id'), 'ZZZ'), x.get('date', '')))
        else:
            sorted_data = sorted(data, key=lambda x: (x.get('date', ''), devs.get(x.get('developer_id'), '')))
        
        # Data rows
        for row_idx, record in enumerate(sorted_data, start=2):
            dev_name = devs.get(record.get('developer_id'), 'Unknown')
            date_val = record.get('date', '')
            commits = self.safe_int(record.get('commits_count'))
            lines_added = self.safe_int(record.get('lines_added'))
            lines_deleted = self.safe_int(record.get('lines_deleted'))
            lines_changed = lines_added + lines_deleted
            files_modified = self.safe_int(record.get('files_modified'))
            churn_score = self.safe_float(record.get('churn_score'))
            coding_secs = self.safe_int(record.get('coding_time_seconds'))
            coding_hhmm = self.format_time_hhmm(coding_secs)
            coding_hours = round(coding_secs / 3600, 2)
            active_mins = round(self.safe_int(record.get('active_coding_seconds')) / 60)
            deep_work_mins = round(self.safe_int(record.get('deep_work_seconds')) / 60)
            focus_ratio = round(self.safe_float(record.get('project_focus_ratio')) * 100, 1)
            context_switches = self.safe_int(record.get('context_switches'))
            reviews = self.safe_int(record.get('review_count'))
            start_time = record.get('start_work_time', '') or ''
            end_time = record.get('end_work_time', '') or ''
            score = self.safe_float(record.get('score'))
            quality_grade = self.get_quality_grade(churn_score)
            updated_at = record.get('updated_at', '') or ''
            
            # Format start/end times
            if start_time and hasattr(start_time, 'strftime'):
                start_time = start_time.strftime('%H:%M')
            if end_time and hasattr(end_time, 'strftime'):
                end_time = end_time.strftime('%H:%M')
            if updated_at and hasattr(updated_at, 'strftime'):
                updated_at = updated_at.strftime('%Y-%m-%d %H:%M:%S')
            
            row_data = [
                dev_name, str(date_val), commits, lines_added, lines_deleted,
                lines_changed, files_modified, round(churn_score, 2), coding_hhmm,
                coding_hours, active_mins, deep_work_mins, focus_ratio,
                context_switches, reviews, str(start_time), str(end_time),
                round(score, 2), quality_grade, str(updated_at)
            ]
            
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.THIN_BORDER
                
                # Alternating row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
            
            # Score conditional formatting (column 18)
            score_cell = ws.cell(row=row_idx, column=18)
            if score >= 80:
                score_cell.fill = self.GREEN_FILL
            elif score >= 50:
                score_cell.fill = self.YELLOW_FILL
            elif score > 0:
                score_cell.fill = self.RED_FILL
            else:
                score_cell.fill = self.GRAY_FILL
        
        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(sorted_data)+1}"
        
        # Column widths
        col_widths = [18, 12, 10, 12, 12, 12, 12, 10, 14, 12, 14, 12, 12, 14, 11, 10, 10, 14, 11, 20]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _create_charts_sheet(self, wb: Workbook, data: List[dict], devs: dict):
        """Create a sheet with embedded charts."""
        ws = wb.create_sheet("Charts & Analysis")
        
        # Prepare daily aggregated data for charts
        daily_scores = {}
        for record in data:
            date_str = str(record.get('date', ''))
            if date_str not in daily_scores:
                daily_scores[date_str] = []
            daily_scores[date_str].append(self.safe_float(record.get('score')))
        
        # Create mini data table for chart
        ws['A1'] = "Daily Performance Data"
        ws['A1'].font = Font(size=14, bold=True)
        
        ws['A3'] = "Date"
        ws['B3'] = "Avg Score"
        
        sorted_dates = sorted(daily_scores.keys())
        for i, date_str in enumerate(sorted_dates, start=4):
            ws[f'A{i}'] = date_str
            ws[f'B{i}'] = round(sum(daily_scores[date_str]) / len(daily_scores[date_str]), 1)
        
        # Create Line Chart
        if len(sorted_dates) > 1:
            chart = LineChart()
            chart.title = "Performance Score Trend"
            chart.style = 10
            chart.y_axis.title = "Score"
            chart.x_axis.title = "Date"
            
            data_ref = Reference(ws, min_col=2, min_row=3, max_row=3 + len(sorted_dates))
            cats_ref = Reference(ws, min_col=1, min_row=4, max_row=3 + len(sorted_dates))
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.height = 12
            chart.width = 20
            
            ws.add_chart(chart, "D3")
        
        # Developer comparison data
        row_offset = max(len(sorted_dates) + 6, 20)
        ws[f'A{row_offset}'] = "Developer Comparison"
        ws[f'A{row_offset}'].font = Font(size=14, bold=True)
        
        ws[f'A{row_offset+2}'] = "Developer"
        ws[f'B{row_offset+2}'] = "Avg Score"
        ws[f'C{row_offset+2}'] = "Total Commits"
        ws[f'D{row_offset+2}'] = "Avg Coding Hrs"
        
        dev_stats = {}
        for record in data:
            dev_id = record.get('developer_id')
            if dev_id not in dev_stats:
                dev_stats[dev_id] = {'scores': [], 'commits': 0, 'coding_secs': []}
            dev_stats[dev_id]['scores'].append(self.safe_float(record.get('score')))
            dev_stats[dev_id]['commits'] += self.safe_int(record.get('commits_count'))
            dev_stats[dev_id]['coding_secs'].append(self.safe_int(record.get('coding_time_seconds')))
        
        for i, (dev_id, stats) in enumerate(dev_stats.items(), start=row_offset+3):
            ws[f'A{i}'] = devs.get(dev_id, 'Unknown')
            ws[f'B{i}'] = round(sum(stats['scores']) / max(len(stats['scores']), 1), 1)
            ws[f'C{i}'] = stats['commits']
            ws[f'D{i}'] = round(sum(stats['coding_secs']) / max(len(stats['coding_secs']), 1) / 3600, 1)
        
        # Create Bar Chart for developers
        if len(dev_stats) > 0:
            bar_chart = BarChart()
            bar_chart.title = "Developer Comparison - Key Metrics"
            bar_chart.style = 10
            
            data_ref = Reference(ws, min_col=2, max_col=4, min_row=row_offset+2, max_row=row_offset+2+len(dev_stats))
            cats_ref = Reference(ws, min_col=1, min_row=row_offset+3, max_row=row_offset+2+len(dev_stats))
            
            bar_chart.add_data(data_ref, titles_from_data=True)
            bar_chart.set_categories(cats_ref)
            bar_chart.height = 12
            bar_chart.width = 20
            
            ws.add_chart(bar_chart, f"F{row_offset}")

    def _create_wakatime_sheet(self, wb: Workbook, data: List[dict], devs: dict):
        """Create sheet with parsed WakaTime data."""
        ws = wb.create_sheet("WakaTime Details")
        
        headers = ["Developer", "Date", "Languages", "Editors", "Projects", "Operating Systems"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = self.HEADER_FILL
        
        row = 2
        for record in data:
            wakatime_str = record.get('wakatime_data')
            if not wakatime_str:
                continue
            
            try:
                wt_data = json.loads(wakatime_str) if isinstance(wakatime_str, str) else wakatime_str
            except json.JSONDecodeError:
                continue
            
            dev_name = devs.get(record.get('developer_id'), 'Unknown')
            date_val = str(record.get('date', ''))
            
            # Extract data
            languages = ", ".join([f"{l.get('name', '?')} ({l.get('percent', 0):.0f}%)" 
                                 for l in wt_data.get('languages', [])[:5]])
            editors = ", ".join([e.get('name', '?') for e in wt_data.get('editors', [])[:3]])
            projects = ", ".join([p.get('name', '?') for p in wt_data.get('projects', [])[:5]])
            os_list = ", ".join([o.get('name', '?') for o in wt_data.get('operating_systems', [])[:3]])
            
            ws.cell(row=row, column=1, value=dev_name)
            ws.cell(row=row, column=2, value=date_val)
            ws.cell(row=row, column=3, value=languages)
            ws.cell(row=row, column=4, value=editors)
            ws.cell(row=row, column=5, value=projects)
            ws.cell(row=row, column=6, value=os_list)
            row += 1
        
        # Column widths
        for i, width in enumerate([18, 12, 40, 25, 40, 25], start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def generate_pdf(
        self,
        metrics_data: List[dict],
        developers_map: dict,
        from_date: str,
        to_date: str,
        include_charts: bool = True,
        include_summary: bool = True,
        group_by_developer: bool = False
    ) -> io.BytesIO:
        """Generate PDF report."""
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Title'],
            fontSize=24,
            spaceAfter=12,
            textColor=colors.HexColor('#3b82f6')
        )
        heading_style = ParagraphStyle(
            'Heading',
            parent=styles['Heading2'],
            fontSize=16,
            spaceBefore=12,
            spaceAfter=6,
            textColor=colors.HexColor('#1e40af')
        )
        
        elements = []
        
        # Cover Page
        elements.append(Spacer(1, 2*inch))
        elements.append(Paragraph("DEVELOPER PERFORMANCE REPORT", title_style))
        elements.append(Spacer(1, 0.5*inch))
        elements.append(Paragraph(f"Date Range: {from_date} to {to_date}", styles['Normal']))
        
        dev_names = list(set([developers_map.get(d.get('developer_id'), 'Unknown') for d in metrics_data]))
        elements.append(Paragraph(f"Developers: {', '.join(dev_names[:5])}{'...' if len(dev_names) > 5 else ''}", styles['Normal']))
        elements.append(Spacer(1, 0.25*inch))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}", styles['Normal']))
        elements.append(PageBreak())
        
        # Executive Summary
        if include_summary and metrics_data:
            elements.append(Paragraph("EXECUTIVE SUMMARY", heading_style))
            elements.append(Spacer(1, 0.25*inch))
            
            total_days = len(metrics_data)
            active_days = len([d for d in metrics_data if d.get('commits_count', 0) > 0])
            total_commits = sum(d.get('commits_count', 0) for d in metrics_data)
            total_lines = sum((d.get('lines_added', 0) or 0) + (d.get('lines_deleted', 0) or 0) for d in metrics_data)
            avg_score = sum(d.get('score', 0) or 0 for d in metrics_data) / max(total_days, 1)
            avg_coding = sum(d.get('coding_time_seconds', 0) or 0 for d in metrics_data) / max(total_days, 1) / 3600
            
            summary_data = [
                ['Metric', 'Value'],
                ['Total Records', str(total_days)],
                ['Active Days', f"{active_days} ({active_days/max(total_days,1)*100:.1f}%)"],
                ['Total Commits', str(total_commits)],
                ['Total Lines Changed', f"{total_lines:,}"],
                ['Average Score', f"{avg_score:.1f}"],
                ['Average Coding Hours', f"{avg_coding:.1f}h"],
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 0.5*inch))
        
        # Charts
        if include_charts and metrics_data:
            elements.append(Paragraph("PERFORMANCE CHARTS", heading_style))
            elements.append(Spacer(1, 0.25*inch))
            
            chart_img = self._generate_trend_chart_image(metrics_data, developers_map)
            if chart_img:
                elements.append(Image(chart_img, width=8*inch, height=4*inch))
                elements.append(Spacer(1, 0.25*inch))
            
            comparison_img = self._generate_comparison_chart_image(metrics_data, developers_map)
            if comparison_img:
                elements.append(Image(comparison_img, width=8*inch, height=4*inch))
            
            elements.append(PageBreak())
        
        # Data Table
        elements.append(Paragraph("DETAILED DATA", heading_style))
        elements.append(Spacer(1, 0.25*inch))
        
        # Simplified table for PDF (fewer columns)
        table_data = [['Developer', 'Date', 'Commits', 'Lines', 'Coding', 'Score', 'Grade']]
        
        sorted_data = sorted(metrics_data, key=lambda x: (x.get('date', ''), developers_map.get(x.get('developer_id'), '')))
        
        for record in sorted_data[:100]:  # Limit rows for PDF
            dev_name = developers_map.get(record.get('developer_id'), 'Unknown')
            date_val = str(record.get('date', ''))[-10:]  # Just the date part
            commits = str(self.safe_int(record.get('commits_count')))
            lines = str(self.safe_int(record.get('lines_added', 0)) + self.safe_int(record.get('lines_deleted', 0)))
            coding = self.format_time_hhmm(self.safe_int(record.get('coding_time_seconds')))
            score = f"{self.safe_float(record.get('score')):.1f}"
            grade = self.get_quality_grade(self.safe_float(record.get('churn_score')))
            
            table_data.append([dev_name[:15], date_val, commits, lines, coding, score, grade])
        
        if len(metrics_data) > 100:
            table_data.append(['...', f'+{len(metrics_data)-100} more', '', '', '', '', ''])
        
        data_table = Table(table_data, colWidths=[1.5*inch, 1*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.7*inch, 0.7*inch])
        
        # Style with conditional formatting on score
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]
        
        # Add conditional formatting for score column
        for i, record in enumerate(sorted_data[:100], start=1):
            score = self.safe_float(record.get('score'))
            if score >= 80:
                table_style.append(('BACKGROUND', (5, i), (5, i), colors.HexColor('#d4edda')))
            elif score >= 50:
                table_style.append(('BACKGROUND', (5, i), (5, i), colors.HexColor('#fff3cd')))
            elif score > 0:
                table_style.append(('BACKGROUND', (5, i), (5, i), colors.HexColor('#f8d7da')))
        
        data_table.setStyle(TableStyle(table_style))
        elements.append(data_table)
        
        doc.build(elements)
        output.seek(0)
        return output

    def _generate_trend_chart_image(self, data: List[dict], devs: dict) -> Optional[io.BytesIO]:
        """Generate performance trend line chart as image."""
        try:
            # Aggregate by date and developer
            dev_data = {}
            for record in data:
                dev_id = record.get('developer_id')
                date_str = str(record.get('date', ''))
                score = self.safe_float(record.get('score'))
                
                if dev_id not in dev_data:
                    dev_data[dev_id] = {}
                dev_data[dev_id][date_str] = score
            
            if not dev_data:
                return None
            
            # Get all dates
            all_dates = sorted(set(d for dates in dev_data.values() for d in dates.keys()))
            
            if len(all_dates) < 2:
                return None
            
            fig, ax = plt.subplots(figsize=(10, 5))
            
            colors_list = ['#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899']
            
            for i, (dev_id, dates) in enumerate(dev_data.items()):
                scores = [dates.get(d, 0) for d in all_dates]
                dev_name = devs.get(dev_id, f'Dev {dev_id}')
                ax.plot(all_dates, scores, marker='o', label=dev_name, 
                       color=colors_list[i % len(colors_list)], linewidth=2)
            
            ax.set_xlabel('Date')
            ax.set_ylabel('Performance Score')
            ax.set_title('Performance Score Trend')
            ax.legend(loc='upper left', fontsize=8)
            ax.grid(True, alpha=0.3)
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            plt.close(fig)
            img_buffer.seek(0)
            return img_buffer
        except Exception as e:
            print(f"Error generating trend chart: {e}")
            return None

    def _generate_comparison_chart_image(self, data: List[dict], devs: dict) -> Optional[io.BytesIO]:
        """Generate developer comparison bar chart as image."""
        try:
            dev_stats = {}
            for record in data:
                dev_id = record.get('developer_id')
                if dev_id not in dev_stats:
                    dev_stats[dev_id] = {'scores': [], 'coding_hrs': []}
                dev_stats[dev_id]['scores'].append(self.safe_float(record.get('score')))
                dev_stats[dev_id]['coding_hrs'].append(self.safe_int(record.get('coding_time_seconds', 0)) / 3600)
            
            if not dev_stats:
                return None
            
            dev_names = [devs.get(d, f'Dev {d}')[:12] for d in dev_stats.keys()]
            avg_scores = [sum(s['scores'])/max(len(s['scores']),1) for s in dev_stats.values()]
            avg_coding = [sum(s['coding_hrs'])/max(len(s['coding_hrs']),1) for s in dev_stats.values()]
            
            fig, ax1 = plt.subplots(figsize=(10, 5))
            
            x = range(len(dev_names))
            width = 0.35
            
            bars1 = ax1.bar([i - width/2 for i in x], avg_scores, width, label='Avg Score', color='#3b82f6')
            ax1.set_ylabel('Average Score', color='#3b82f6')
            ax1.tick_params(axis='y', labelcolor='#3b82f6')
            ax1.set_ylim(0, 100)
            
            ax2 = ax1.twinx()
            bars2 = ax2.bar([i + width/2 for i in x], avg_coding, width, label='Avg Coding Hours', color='#10b981')
            ax2.set_ylabel('Average Coding Hours', color='#10b981')
            ax2.tick_params(axis='y', labelcolor='#10b981')
            ax2.set_ylim(0, max(avg_coding) * 1.2 if avg_coding else 10)
            
            ax1.set_xlabel('Developer')
            ax1.set_title('Developer Comparison - Key Metrics')
            ax1.set_xticks(x)
            ax1.set_xticklabels(dev_names, rotation=45, ha='right')
            
            fig.legend(loc='upper right', bbox_to_anchor=(0.95, 0.95))
            plt.tight_layout()
            
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            plt.close(fig)
            img_buffer.seek(0)
            return img_buffer
        except Exception as e:
            print(f"Error generating comparison chart: {e}")
            return None
